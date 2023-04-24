import mysql.connector
from openpyxl import Workbook, load_workbook

file_path = "catalogue.xlsx"

def get_data():
    mydb = 0
    try:
        mydb = mysql.connector.connect(
            host="",
            user="anuarAdmin",
            password="",
            database="voltmandb"
        )
    except:
        print("An exception occured!!!")

    cursor = mydb.cursor()

    cursor.execute("SELECT name, type, applicability, size_group, polarity, voltage, capacity, startup, terminals, hold, manufacturer, country, brand FROM car_battery_catalogue;")
    data = cursor.fetchall()

    cursor.close()
    mydb.close()

    return data

def create_excel(data):
    wb = Workbook()
    ws = wb.create_sheet("data")

    ws.append(('name', 'type', 'applicability', 'size_group', 'polarity', 'voltage', 'capacity', 'startup', 'terminals', 'hold', 'manufacturer', 'country', 'brand'))

    for i in data:
        ws.append(i)

    wb.save(file_path)

def main():
    data = get_data()

    create_excel(data)

if __name__ == "__main__":
    main()

