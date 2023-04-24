import mysql.connector
from openpyxl import Workbook, load_workbook

file_path = "file.xlsx"

def get_data():

    mydb = 0
    try:
        mydb = mysql.connector.connect(
            host="",
            user="anuarAdmin",
            password="",
            database="voltmandb"
        )
    except Error as e:
        print(e)

    cursor = mydb.cursor()

    all_list = []
    distinct_list = [("good_name", "name")]

    cursor.execute("SELECT unique_good_name, dealer_good_name, rest, warehouse FROM goods GROUP BY 1c_good_id")
    data = cursor.fetchall()
    for i in data:
        distinct_list.append(i)

    cursor.execute("SELECT dealer_good_name, rest FROM goods")
    data = cursor.fetchall()    
    for i in data:
        all_list.append(i)

    print("all", len(all_list))
    print("distinct", len(distinct_list))

    mydb.disconnect()

    return distinct_list

def excel_file(data):
    wb = Workbook()
    ws =wb.create_sheet("distinct")

    for i in data:
        ws.append(i)

    wb.save(file_path)

def main():    
    print("Main")

    data = get_data()

    excel_file(data)

if __name__ == "__main__":
    main()